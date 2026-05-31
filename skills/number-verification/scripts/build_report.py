#!/usr/bin/env python3
"""
build_report.py — verified JSON → Excel verification report (6 + 1 columns)

Usage:
    python3 build_report.py <verified_json> [<output_xlsx>]

Input JSON (extends extract_numbers.py output with `verification` per candidate):
{
  "source_file": "report.docx",
  "candidates": [
    {
      "id": 1,
      "value": "1.2亿台",
      "raw": "1.2亿",
      "context": "...",
      "location": "page 2, paragraph 3",
      "verification": {
        "web_value": "9000 万台 (2023)",
        "web_source": "https://... | OICA 国际汽车制造商组织",
        "sanity_calc": "全球 80 亿人 / 16 亿家庭 · 保有量 ~10 亿 · 更换周期 10 年 → 年销 ~1 亿量级",
        "verdict": "⚠ 存疑",
        "note": "文档值偏高 ~33%；可能 includes 商用车 + 摩托车口径"
      }
    }
  ]
}

If `verification` field is missing on a candidate, it'll be left blank
(reviewer can fill in afterwards).
"""

import sys
import os
import json
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl 未安装。请 `pip install openpyxl`")

FONT_NAME = "Arial"


HEADERS = [
    ("数字值",       "value",       14),
    ("文档位置",     "location",    22),
    ("文档原文上下文","context",     56),
    ("检索验证",     "web_value",   22),
    ("检索 source",  "web_source",  40),
    ("常识推算",     "sanity_calc", 56),
    ("判定结果",     "verdict",     12),
    ("备注",         "note",        40),
]

VERDICT_FILL = {
    "✓": PatternFill("solid", fgColor="E2EFDA"),  # green
    "⚠": PatternFill("solid", fgColor="FFF2CC"),  # yellow
    "✗": PatternFill("solid", fgColor="F8CBAD"),  # red
}
VERDICT_RANK = {"✗": 0, "⚠": 1, "✓": 2, "—": 3, "": 4}


def verdict_key(c):
    v = (c.get("verification") or {}).get("verdict", "")
    first = (v or "").strip()[:1]
    return VERDICT_RANK.get(first, 5)


def build_workbook(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Number Verification"

    # ---- summary block at top ----
    ws["A1"] = "Number Verification Report"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    src = data.get("source_file", "—")
    extracted = data.get("extracted_at", "—")
    cands = data.get("candidates", [])

    counts = {"✓": 0, "⚠": 0, "✗": 0, "—": 0, "_blank": 0}
    for c in cands:
        v = (c.get("verification") or {}).get("verdict", "").strip()
        first = v[:1] if v else ""
        if first in counts:
            counts[first] += 1
        elif v:
            counts["—"] += 1
        else:
            counts["_blank"] += 1

    ws["A2"] = f"Source: {src}   Extracted: {extracted}   Total: {len(cands)}   ✓ {counts['✓']}   ⚠ {counts['⚠']}   ✗ {counts['✗']}   未填: {counts['_blank']}"
    ws["A2"].font = Font(name=FONT_NAME, size=10, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20

    # ---- column headers (row 4) ----
    header_row = 4
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(top=thin, right=thin, bottom=thin, left=thin)

    for i, (label, _, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=header_row, column=i, value=label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[header_row].height = 32

    # ---- data rows ----
    body_font = Font(name=FONT_NAME, size=10)
    body_align = Alignment(vertical="top", wrap_text=True)
    sorted_cands = sorted(cands, key=verdict_key)

    for idx, c in enumerate(sorted_cands, start=header_row + 1):
        v = c.get("verification") or {}
        verdict_text = (v.get("verdict") or "").strip()
        verdict_head = verdict_text[:1]

        row_vals = {
            "value":       c.get("value", ""),
            "location":    c.get("location", ""),
            "context":     c.get("context", ""),
            "web_value":   v.get("web_value", ""),
            "web_source":  v.get("web_source", ""),
            "sanity_calc": v.get("sanity_calc", ""),
            "verdict":     verdict_text,
            "note":        v.get("note", ""),
        }
        for col_idx, (_, key, _) in enumerate(HEADERS, 1):
            cell = ws.cell(row=idx, column=col_idx, value=row_vals.get(key, ""))
            cell.font = body_font
            cell.alignment = body_align
            cell.border = border
            if key == "verdict" and verdict_head in VERDICT_FILL:
                cell.fill = VERDICT_FILL[verdict_head]
                cell.font = Font(name=FONT_NAME, size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[idx].height = 50

    # freeze header + summary
    ws.freeze_panes = f"A{header_row + 1}"

    # ---- legend sheet ----
    legend = wb.create_sheet("Legend")
    legend.append(["列", "含义"])
    legend["A1"].font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
    legend["B1"].font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
    legend["A1"].fill = header_fill
    legend["B1"].fill = header_fill
    for label, key, _ in HEADERS:
        legend.append([label, key])
    legend.append([])
    legend.append(["判定符号", "含义 / 触发条件"])
    legend["A" + str(legend.max_row)].font = Font(name=FONT_NAME, bold=True)
    legend["B" + str(legend.max_row)].font = Font(name=FONT_NAME, bold=True)
    legend.append(["✓ 正确",  "检索值与文档值差异 < 10% 且推算量级吻合"])
    legend.append(["⚠ 存疑",  "差异 10%–50%，或检索失败但推算量级勉强吻合"])
    legend.append(["✗ 错误",  "差异 > 50%，或推算量级明显不符（数量级错）"])
    legend.append(["— 非数据型", "页码 / 序号 / 章节号等非数据数字"])
    # apply Arial to all body cells in legend sheet
    for row in legend.iter_rows(min_row=2):
        for cell in row:
            if cell.font.name != FONT_NAME:
                cell.font = Font(name=FONT_NAME, size=10,
                                 bold=cell.font.bold or False)
    legend.column_dimensions["A"].width = 18
    legend.column_dimensions["B"].width = 70

    return wb


def main():
    parser = argparse.ArgumentParser(description="Build verification Excel report.")
    parser.add_argument("input_json", help="verified JSON path")
    parser.add_argument("output_xlsx", nargs="?", help="output xlsx path (default: alongside input)")
    args = parser.parse_args()

    if not os.path.isfile(args.input_json):
        sys.exit(f"json not found: {args.input_json}")

    with open(args.input_json, "r", encoding="utf-8") as f:
        data = json.load(f)

    if args.output_xlsx:
        out_path = args.output_xlsx
    else:
        base = os.path.splitext(os.path.basename(data.get("source_file") or args.input_json))[0]
        out_dir = os.path.dirname(os.path.abspath(args.input_json))
        stamp = datetime.now().strftime("%Y%m%d-%H%M")
        out_path = os.path.join(out_dir, f"{base}_verification_{stamp}.xlsx")

    wb = build_workbook(data)
    wb.save(out_path)
    print(f"✓ wrote report → {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
